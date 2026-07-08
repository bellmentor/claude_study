"""매입금 수집 Web UI 서버.

각 도매처 모듈을 subprocess(python -m dome_site.<slug>.main)로 실행하고,
stdout 출력을 WebSocket으로 브라우저에 실시간 전달한다.
"""

from __future__ import annotations

import asyncio
import subprocess
import sys
import threading
from pathlib import Path
from typing import Any

# 서버 콘솔 출력을 UTF-8 로 고정한다.
# (Windows cp949 콘솔/파이프에서 한글·기호가 깨지거나 UnicodeEncodeError 로 죽는 것을 방지)
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]
    except Exception:
        pass

import openpyxl
import pandas as pd
from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request

# ── 경로 설정 ──────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
APP_DIR = Path(__file__).resolve().parent
ACCOUNT_XLSX = ROOT / "계정정보.xlsx"
SITE_LIST_TXT = ROOT / "dome_site" / "도매사이트_폴더_이름.txt"
SUMMARY_XLSX = ROOT / "dome_site" / "도매_매입금.xlsx"
PYTHON = sys.executable

# ── 정산엑셀 관리 ──────────────────────────────────────────
# 프로그램이 계속 참조할 정산용 엑셀을 명명된 슬롯에 업로드/교체/삭제한다.
# 파일은 슬롯 key 로 고정 저장(정산엑셀/<key>.xlsx)하여 안정 경로로 참조 가능.
# 슬롯 추가 = 아래 SETTLEMENT_SLOTS 에 {"key","name"} 한 줄 추가.
SETTLEMENT_DIR = ROOT / "정산엑셀"
SETTLEMENT_MANIFEST = SETTLEMENT_DIR / "_manifest.json"
SETTLEMENT_SLOTS: list[dict[str, str]] = [
    {"key": "total", "name": "전체 정산용 엑셀"},
]
ALLOWED_EXCEL_EXT = (".xlsx", ".xls")

# 에이준줄눈: 폴더째 업로드받아 그 안의 엑셀들로 매입금을 계산할 작업폴더.
# (브라우저가 폴더 절대경로를 못 주므로, 선택 폴더의 엑셀 파일들을 여기로 복사한다.)
AEJULNUN_DIR = SETTLEMENT_DIR / "aejulnun"
AEJULNUN_META = AEJULNUN_DIR / "_meta.json"

# BearB2B: 루트 ./BearB2B 폴더에서 자체 크롤링/계산 코드를 둔다(스캐폴드 단계).
BEARB2B_DIR = ROOT / "BearB2B"

# ── 로그 버퍼 + WebSocket 브로드캐스터 ─────────────────────
# 로그는 HTTP 폴링(상태 조회)으로 받는 것을 1차 채널로 한다. WebSocket 은
# 리로드/늦은 연결 시 메시지를 놓치므로, 버퍼에 쌓아두고 폴링으로 재생한다.
_logs: list[str] = []
_ws_clients: set[WebSocket] = set()
_collect_status: dict[str, dict[str, Any]] = {}
_collect_running = False
_main_loop: asyncio.AbstractEventLoop | None = None

# BearB2B(고도몰) 수집 상태 (매입금 수집과 별개 흐름)
_bearb2b_status: dict[str, Any] = {}
_bearb2b_running = False


from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    global _main_loop
    _main_loop = asyncio.get_running_loop()
    yield


# ── FastAPI 앱 ─────────────────────────────────────────────
app = FastAPI(title="매입금 수집", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=APP_DIR / "static"), name="static")
templates = Jinja2Templates(directory=APP_DIR / "templates")


async def broadcast_log(message: str) -> None:
    """연결된 모든 WebSocket 클라이언트에 로그 메시지를 전송한다."""
    dead: set[WebSocket] = set()
    for ws in _ws_clients:
        try:
            await ws.send_text(message)
        except Exception:
            dead.add(ws)
    _ws_clients.difference_update(dead)


def send_log(message: str) -> None:
    """어느 스레드에서든 호출 가능한 로그 전송. 버퍼에 쌓고 WebSocket 으로도 보낸다."""
    msg = message.strip()
    if not msg:
        return
    _logs.append(msg)  # 폴링으로 재생할 버퍼 (1차 채널)
    if _main_loop:  # 연결돼 있으면 실시간 전송 (보조 채널)
        asyncio.run_coroutine_threadsafe(broadcast_log(msg), _main_loop)


# ── 사이트별 진행상황 주의 문구 ────────────────────────────
# 진행상황 컬럼에 '항상' 표시할 주의 문구(slug 기준). 구현상 한계/미완성을 눈에 띄게 남긴다.
SITE_NOTES: dict[str, str] = {
    "coms": "페이지 넘어갈시 코드수정필요",  # 컴스마트: 주문내역 단일 페이지만 크롤링함
    "sic": "조회 6개월 범위 (6달 지난 달은 앞부분 누락 주의)",  # 식자재코리아: '6개월' 버튼으로 조회
    "cokorea": "조회 6개월 범위 (6달 지난 달은 누락 주의)",  # 코코리아: '6개월' 버튼으로 조회
}


# ── 도매사이트 목록 파싱 ───────────────────────────────────
def load_site_list() -> list[dict[str, str]]:
    """도매사이트_폴더_이름.txt 를 파싱하여 [{name, slug, note}] 반환."""
    sites: list[dict[str, str]] = []
    text = SITE_LIST_TXT.read_text(encoding="utf-8")
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("----"):
            break
        if ":" in line:
            name, slug = line.split(":", 1)
            name, slug = name.strip(), slug.strip()
            if slug:
                sites.append({"name": name, "slug": slug, "note": SITE_NOTES.get(slug, "")})
    return sites


# ── 정산엑셀 헬퍼 ──────────────────────────────────────────
def _load_settlement_manifest() -> dict[str, dict[str, str]]:
    """정산엑셀/_manifest.json 을 읽어 반환한다. 없으면 빈 dict."""
    if not SETTLEMENT_MANIFEST.exists():
        return {}
    try:
        import json
        return json.loads(SETTLEMENT_MANIFEST.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_settlement_manifest(manifest: dict[str, dict[str, str]]) -> None:
    """정산엑셀/_manifest.json 에 저장한다."""
    import json
    SETTLEMENT_DIR.mkdir(exist_ok=True)
    SETTLEMENT_MANIFEST.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def settlement_file_path(key: str) -> Path | None:
    """슬롯 key 의 업로드된 정산엑셀 경로를 반환한다. 없으면 None.

    향후 프로그램(정산 계산 로직 등)이 이 헬퍼로 업로드된 엑셀을 읽어 사용한다.
    """
    if key not in {s["key"] for s in SETTLEMENT_SLOTS}:
        return None
    path = SETTLEMENT_DIR / f"{key}.xlsx"
    return path if path.exists() else None


def _settlement_status() -> list[dict[str, Any]]:
    """SETTLEMENT_SLOTS 순회하며 각 슬롯의 업로드 상태를 만든다."""
    manifest = _load_settlement_manifest()
    slots: list[dict[str, Any]] = []
    for slot in SETTLEMENT_SLOTS:
        key = slot["key"]
        path = SETTLEMENT_DIR / f"{key}.xlsx"
        uploaded = path.exists()
        meta = manifest.get(key, {})
        slots.append({
            "key": key,
            "name": slot["name"],
            "uploaded": uploaded,
            "filename": meta.get("original", "") if uploaded else "",
            "uploaded_at": meta.get("uploaded_at", "") if uploaded else "",
        })
    return slots


# ── 에이준줄눈 폴더 헬퍼 ───────────────────────────────────
def aejulnun_files() -> list[Path]:
    """업로드된 에이준줄눈 작업폴더의 엑셀 파일 경로 목록을 반환한다.

    향후 에이준줄눈 매입금 계산 로직이 이 헬퍼로 폴더 안 엑셀들을 읽어 계산한다.
    """
    if not AEJULNUN_DIR.exists():
        return []
    return sorted(
        f for f in AEJULNUN_DIR.iterdir()
        if f.is_file()
        and f.suffix.lower() in ALLOWED_EXCEL_EXT
        and not f.name.startswith("~")
    )


def _load_aejulnun_meta() -> dict[str, Any]:
    """에이준줄눈 작업폴더 메타(_meta.json)를 읽어 반환한다. 없으면 빈 dict."""
    if not AEJULNUN_META.exists():
        return {}
    try:
        import json
        return json.loads(AEJULNUN_META.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_aejulnun_meta(meta: dict[str, Any]) -> None:
    """에이준줄눈 작업폴더 메타를 저장한다."""
    import json
    AEJULNUN_DIR.mkdir(parents=True, exist_ok=True)
    AEJULNUN_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def _wipe_aejulnun_dir() -> None:
    """에이준줄눈 작업폴더의 파일을 모두 지운다(폴더 자체는 유지)."""
    if not AEJULNUN_DIR.exists():
        return
    for f in AEJULNUN_DIR.iterdir():
        if f.is_file():
            try:
                f.unlink()
            except Exception:
                pass


def _aejulnun_status() -> dict[str, Any]:
    """에이준줄눈 작업폴더 상태(선택폴더명/시각/엑셀목록)를 만든다."""
    meta = _load_aejulnun_meta()
    files = [{"name": f.name, "size": f.stat().st_size} for f in aejulnun_files()]
    return {
        "folder": meta.get("folder", ""),
        "uploaded_at": meta.get("uploaded_at", ""),
        "count": len(files),
        "files": files,
    }


# ── 사이트별 subprocess 실행 ──────────────────────────────
def _run_site_subprocess(slug: str, name: str, year: int, month: int) -> None:
    """별도 스레드에서 python -m dome_site.<slug>.main 을 subprocess로 실행한다."""
    _collect_status[slug] = {"name": name, "status": "실행 중...", "amount": "", "error": ""}
    send_log(f"[{name}] 수집 시작 ({year}년 {month}월)")

    try:
        env = {**__import__("os").environ, "WEBUI": "1", "PYTHONIOENCODING": "utf-8"}
        proc = subprocess.Popen(
            [PYTHON, "-m", f"dome_site.{slug}.main", str(year), str(month)],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            cwd=str(ROOT),
            env=env,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

        for line in proc.stdout:
            line = line.rstrip()
            if line:
                # 서버를 띄운 터미널에도 그대로 출력 (도매처 상세 로그를 콘솔에서 확인)
                print(line, flush=True)
                # 브라우저 로그 박스로 전송
                send_log(line)

        proc.wait()

        if proc.returncode == 0:
            _collect_status[slug]["status"] = "완료"
            send_log(f"[{name}] 수집 완료")
        else:
            _collect_status[slug]["status"] = "오류"
            _collect_status[slug]["error"] = f"오류 (종료코드: {proc.returncode})"
            send_log(f"[{name}] 오류 발생 (종료코드: {proc.returncode})")

    except Exception as e:
        _collect_status[slug]["status"] = "오류"
        _collect_status[slug]["error"] = f"{e}"
        send_log(f"[{name}] 실행 오류: {e}")


def _run_all_sites(sites: list[dict], year: int, month: int) -> None:
    """모든 선택된 사이트를 순차 실행한다. 별도 스레드에서 호출."""
    global _collect_running
    try:
        for site in sites:
            _run_site_subprocess(site["slug"], site["name"], year, month)

        # 완료 후 매입금 읽기
        _read_amounts(year, month)
        send_log("=== 모든 수집 완료 ===")
    finally:
        _collect_running = False


def _read_amounts(year: int, month: int) -> None:
    """도매_매입금.xlsx 에서 해당 월의 매입금을 읽어 status에 반영한다."""
    if not SUMMARY_XLSX.exists():
        return
    try:
        df = pd.read_excel(SUMMARY_XLSX)
        yy, mm = str(year)[2:], f"{month:02d}"
        month_label = f"{yy}년{mm}월"
        # slug → 한글이름 매핑
        site_list = load_site_list()
        slug_to_name = {s["slug"]: s["name"] for s in site_list}
        for slug, info in _collect_status.items():
            korean_name = slug_to_name.get(slug, info["name"])
            row = df[(df["몇월"] == month_label) & (df["도매사이트"] == korean_name)]
            if not row.empty:
                amount = int(row.iloc[0]["매입금"])
                _collect_status[slug]["amount"] = f"{amount:,}"
    except Exception:
        pass


# ── BearB2B subprocess 실행 ───────────────────────────────
def _run_bearb2b(year: int, month: int) -> None:
    """별도 스레드에서 python -m BearB2B.main 을 subprocess로 실행한다."""
    global _bearb2b_running
    _bearb2b_status.update({"status": "실행 중...", "amount": "", "error": ""})
    send_log(f"[베어B2B] 수집 시작 ({year}년 {month}월)")

    try:
        env = {**__import__("os").environ, "WEBUI": "1", "PYTHONIOENCODING": "utf-8"}
        proc = subprocess.Popen(
            [PYTHON, "-m", "BearB2B.main", str(year), str(month)],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            cwd=str(ROOT),
            env=env,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

        for line in proc.stdout:
            line = line.rstrip()
            if line:
                print(line, flush=True)
                send_log(line)

        proc.wait()

        if proc.returncode == 0:
            _bearb2b_status["status"] = "완료"
            _bearb2b_status["amount"] = _read_bearb2b_amount(year, month)
            send_log("[베어B2B] 수집 완료")
        else:
            _bearb2b_status["status"] = "오류"
            _bearb2b_status["error"] = f"오류 (종료코드: {proc.returncode})"
            send_log(f"[베어B2B] 오류 발생 (종료코드: {proc.returncode})")

    except Exception as e:
        _bearb2b_status["status"] = "오류"
        _bearb2b_status["error"] = f"{e}"
        send_log(f"[베어B2B] 실행 오류: {e}")
    finally:
        _bearb2b_running = False


def _read_bearb2b_amount(year: int, month: int) -> str:
    """도매_매입금.xlsx 에서 해당 월 베어B2B 매입금을 읽는다. 없으면 빈 문자열."""
    if not SUMMARY_XLSX.exists():
        return ""
    try:
        df = pd.read_excel(SUMMARY_XLSX)
        month_label = f"{str(year)[2:]}년{month:02d}월"
        row = df[(df["몇월"] == month_label) & (df["도매사이트"] == "베어B2B")]
        if not row.empty:
            return f"{int(row.iloc[0]['매입금']):,}"
    except Exception:
        pass
    return ""


# ── 페이지 ─────────────────────────────────────────────────
def _static_version() -> int:
    """정적 파일(app.js/style.css) 최신 수정시각 → 캐시 무력화 버전.

    <script src="/static/app.js?v=..."> 형태로 붙여, 파일이 바뀌면 URL 도 바뀌게 한다.
    브라우저 캐시 때문에 "코드 고쳤는데 그대로"인 착시(Ctrl+F5 필요)를 없앤다.
    """
    static_dir = APP_DIR / "static"
    try:
        return max(int((static_dir / f).stat().st_mtime) for f in ("app.js", "style.css"))
    except OSError:
        return 0


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """메인 페이지."""
    sites = load_site_list()
    return templates.TemplateResponse(
        request, "index.html", {"sites": sites, "v": _static_version()}
    )


# ── 계정 API ──────────────────────────────────────────────
@app.get("/api/accounts")
async def get_accounts():
    """계정정보.xlsx 에서 계정 목록을 읽어 반환한다."""
    if not ACCOUNT_XLSX.exists():
        return {"accounts": [], "error": "계정정보.xlsx 파일이 없습니다"}

    wb = openpyxl.load_workbook(ACCOUNT_XLSX, data_only=True)
    ws = wb["Sheet1"]
    accounts = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and row[0]:
            accounts.append({
                "site": str(row[0]),
                "user_id": str(row[1]) if row[1] else "",
                "password": str(row[2]) if row[2] else "",
            })
    wb.close()
    return {"accounts": accounts}


@app.put("/api/accounts")
async def put_accounts(request: Request):
    """계정정보를 수정하여 xlsx 에 저장한다."""
    if not ACCOUNT_XLSX.exists():
        return {"error": "계정정보.xlsx 파일이 없습니다. 직접 생성해주세요."}

    body = await request.json()
    accounts = body.get("accounts", [])

    wb = openpyxl.load_workbook(ACCOUNT_XLSX)
    ws = wb["Sheet1"]
    for r_idx in range(1, ws.max_row + 1):
        for c_idx in range(1, 4):
            ws.cell(row=r_idx, column=c_idx, value=None)

    for i, acc in enumerate(accounts, start=1):
        ws.cell(row=i, column=1, value=acc.get("site", ""))
        ws.cell(row=i, column=2, value=acc.get("user_id", ""))
        ws.cell(row=i, column=3, value=acc.get("password", ""))

    wb.save(ACCOUNT_XLSX)
    wb.close()
    return {"ok": True}


# ── 정산엑셀 API ──────────────────────────────────────────
@app.get("/api/settlement")
async def get_settlement():
    """정산엑셀 슬롯 목록 + 각 슬롯의 업로드 상태를 반환한다."""
    return {"slots": _settlement_status()}


@app.post("/api/settlement/upload/{key}")
async def upload_settlement(key: str, request: Request):
    """정산엑셀을 슬롯 key 로 업로드한다(raw body). 저장 후 갱신된 슬롯 상태 반환.

    python-multipart 없이 동작하도록 파일 바이트를 요청 body 로 직접 받는다.
    원본 파일명은 X-Filename 헤더로 전달한다(encodeURIComponent → 서버에서 unquote).
    ※ 쿼리파라미터 대신 헤더를 쓰는 이유: 파일명이 URL 에 들어가면 uvicorn 접속 로그에
      퍼센트 인코딩(%EC%9B%94…)으로 찍혀 콘솔에서 한글이 안 읽힌다. 헤더는 접속 로그에
      안 남으므로 경로가 깨끗해지고, 아래에서 읽기 쉬운 한글 로그를 따로 찍는다.
    """
    import urllib.parse
    filename = urllib.parse.unquote(request.headers.get("x-filename", ""), encoding="utf-8")

    valid_keys = {s["key"] for s in SETTLEMENT_SLOTS}
    if key not in valid_keys:
        return {"error": f"알 수 없는 슬롯: {key}"}

    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXCEL_EXT:
        return {"error": "엑셀 파일(.xlsx, .xls)만 업로드할 수 있습니다"}

    data = await request.body()
    if not data:
        return {"error": "빈 파일입니다"}

    SETTLEMENT_DIR.mkdir(exist_ok=True)
    dest = SETTLEMENT_DIR / f"{key}.xlsx"
    try:
        dest.write_bytes(data)
    except PermissionError:
        return {"error": "파일이 열려있어 저장할 수 없습니다. 엑셀에서 닫고 다시 시도하세요."}

    import datetime
    manifest = _load_settlement_manifest()
    manifest[key] = {
        "original": filename,
        "uploaded_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _save_settlement_manifest(manifest)

    # 서버 콘솔에도 읽기 쉬운 한글 로그(stdout 은 상단에서 UTF-8 로 고정됨)
    print(f"[정산엑셀] '{filename}' 업로드됨 → {key} ({len(data):,} bytes)", flush=True)
    send_log(f"[정산엑셀] '{filename}' 업로드됨 → {key}")
    return {"ok": True, "slots": _settlement_status()}


@app.delete("/api/settlement/{key}")
async def delete_settlement(key: str):
    """슬롯 key 의 정산엑셀을 삭제한다."""
    valid_keys = {s["key"] for s in SETTLEMENT_SLOTS}
    if key not in valid_keys:
        return {"error": f"알 수 없는 슬롯: {key}"}

    path = SETTLEMENT_DIR / f"{key}.xlsx"
    if path.exists():
        try:
            path.unlink()
        except PermissionError:
            return {"error": "파일이 열려있어 삭제할 수 없습니다. 엑셀에서 닫고 다시 시도하세요."}

    manifest = _load_settlement_manifest()
    manifest.pop(key, None)
    _save_settlement_manifest(manifest)

    print(f"[정산엑셀] {key} 삭제됨", flush=True)
    send_log(f"[정산엑셀] {key} 삭제됨")
    return {"ok": True, "slots": _settlement_status()}


# ── BearB2B API ────────────────────────────────────────────
@app.get("/api/bearb2b")
async def get_bearb2b():
    """BearB2B 탭 상태(실행 여부/마지막 결과)를 반환한다."""
    return {"running": _bearb2b_running, "status": _bearb2b_status}


@app.post("/api/bearb2b/run")
async def run_bearb2b(request: Request):
    """BearB2B(고도몰) 매입금 수집을 시작한다."""
    global _bearb2b_running
    if _bearb2b_running:
        return {"error": "이미 베어B2B 수집이 진행 중입니다"}
    if _collect_running:
        return {"error": "매입금 수집이 진행 중입니다. 끝난 뒤 실행해주세요"}

    body = await request.json()
    try:
        year, month = int(body.get("year")), int(body.get("month"))
    except (TypeError, ValueError):
        return {"error": "년/월을 올바르게 선택해주세요"}

    _bearb2b_running = True
    thread = threading.Thread(target=_run_bearb2b, args=(year, month), daemon=True)
    thread.start()
    return {"ok": True, "message": "베어B2B 수집을 시작했습니다"}


@app.get("/api/bearb2b/status")
async def bearb2b_status(since: int = 0):
    """베어B2B 진행상황 + since 이후 새 로그를 반환한다 (HTTP 폴링)."""
    new_logs = _logs[since:] if since < len(_logs) else []
    return {
        "running": _bearb2b_running,
        "status": _bearb2b_status,
        "logs": new_logs,
        "log_count": len(_logs),
    }


# ── 에이준줄눈 폴더 API ────────────────────────────────────
@app.get("/api/aejulnun")
async def get_aejulnun():
    """에이준줄눈 작업폴더 상태(선택폴더명/엑셀목록)를 반환한다."""
    return _aejulnun_status()


@app.post("/api/aejulnun/clear")
async def clear_aejulnun(request: Request):
    """새 폴더 업로드 시작: 기존 작업폴더를 비우고 메타(선택폴더명/시각)를 초기화한다.

    선택한 폴더명은 X-Folder 헤더로 전달한다(encodeURIComponent → 서버 unquote).
    """
    import urllib.parse, datetime
    folder = urllib.parse.unquote(request.headers.get("x-folder", ""), encoding="utf-8")

    _wipe_aejulnun_dir()
    _save_aejulnun_meta({
        "folder": folder,
        "uploaded_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    print(f"[에이준줄눈] 폴더 선택: '{folder}' (기존 작업폴더 비움)", flush=True)
    send_log(f"[에이준줄눈] 폴더 선택: '{folder}'")
    return {"ok": True}


@app.post("/api/aejulnun/upload")
async def upload_aejulnun(request: Request):
    """폴더 안의 엑셀 파일 1개를 작업폴더로 업로드한다(raw body).

    원본 상대경로는 X-Filename 헤더로 전달한다. 경로 구분자는 __ 로 평탄화해 저장.
    엑셀(.xlsx/.xls)이 아니면 무시(폴더에 섞인 다른 파일 방어).
    """
    import urllib.parse
    relpath = urllib.parse.unquote(request.headers.get("x-filename", ""), encoding="utf-8")

    ext = Path(relpath).suffix.lower()
    if ext not in ALLOWED_EXCEL_EXT:
        return {"skipped": relpath}

    # "폴더/하위/a.xlsx" → 최상위 폴더명 떼고 나머지 경로구분자 평탄화
    parts = relpath.replace("\\", "/").split("/")
    inner = parts[1:] if len(parts) > 1 else parts
    safe = "__".join(inner) or Path(relpath).name

    data = await request.body()
    if not data:
        return {"error": "빈 파일입니다"}

    AEJULNUN_DIR.mkdir(parents=True, exist_ok=True)
    dest = AEJULNUN_DIR / safe
    try:
        dest.write_bytes(data)
    except PermissionError:
        return {"error": f"파일이 열려있어 저장 실패: {safe}"}
    return {"ok": True, "saved": safe}


# ── 에이준줄눈 계산 헬퍼 ───────────────────────────────────
def _ajn_to_int(v) -> int:
    """셀 값을 정수로. '3,000'/3000.0/'-4000' 등 처리, 빈 값/비숫자는 0."""
    import re
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0
    if isinstance(v, bool):
        return 0
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(round(v))
    s = str(v).strip().replace(",", "")
    m = re.match(r"^-?\d+", s)
    return int(m.group()) if m else 0


def _leading_mmdd(name: str) -> int:
    """파일명 앞 4자리 숫자(MMDD)를 정수로. 없으면 큰 값(뒤로 정렬)."""
    import re
    m = re.match(r"\s*(\d{4})", name)
    return int(m.group(1)) if m else 99999


def _parse_order_form(path: Path) -> dict[str, Any]:
    """발주서 엑셀 1개를 파싱한다.

    반환: {orders:[{수령인, 송장, amount}], net:요약행순액|None, total:요약행합계|None}
      - '상품명' 열 오른쪽 ~ '택배사명(발송방법)' 열 왼쪽 = 가격 블록(배송비+상품가+보수제).
      - 요약행 = 상품명 셀에 '=' 있는 행('순액+부가세=합계'). 숫자 파싱해 net(첫값)/total(끝값).
        ★ 요약행에서 파싱을 멈춘다(break) — 그 아래 '2/28 미수금 39,050 포함' 같은 메모 행을
          주문으로 잘못 합산하지 않기 위해.
      - 주문행 = 요약행 위쪽의, 상품명 비지 않은 행('반품' 포함). 가격블록 합 = 그 주문 금액.
      - 송장번호는 첫/마지막 파일 보정(월 소속 판정)에 쓴다.
    """
    import re
    raw = pd.read_excel(path, header=None, dtype=object)
    ncol = raw.shape[1]
    header = ["" if pd.isna(raw.iat[0, c]) else str(raw.iat[0, c]).strip() for c in range(ncol)]

    def col_of(*kws):
        for i, h in enumerate(header):
            for kw in kws:
                if kw in h:
                    return i
        return None

    name_col = col_of("상품명")
    ship_col = col_of("택배사")
    recv_col = col_of("수령인명", "수령인", "받는분", "받는사람", "수취인")
    inv_col = col_of("송장")
    if name_col is None:
        raise ValueError(f"'상품명' 열을 찾지 못함: {header[:10]}")
    block_end = ship_col if (ship_col is not None and ship_col > name_col) else min(name_col + 8, ncol)
    block_cols = list(range(name_col + 1, block_end))

    orders: list[dict[str, Any]] = []
    net = total = None
    for r in range(1, raw.shape[0]):
        pv = raw.iat[r, name_col]
        pname = "" if pd.isna(pv) else str(pv).strip()
        if not pname:
            continue
        if "=" in pname:  # 요약행 → 여기서 종료(아래는 메모/주석 행)
            nums = [int(x.replace(",", "")) for x in re.findall(r"-?[\d,]+", pname)
                    if x.replace(",", "").lstrip("-").isdigit()]
            if len(nums) >= 3:
                net, total = nums[0], nums[-1]
            elif nums:
                total = nums[-1]
            break
        amt = sum(_ajn_to_int(raw.iat[r, c]) for c in block_cols)
        recv = ""
        if recv_col is not None and recv_col < ncol:
            rv = raw.iat[r, recv_col]
            recv = "" if pd.isna(rv) else str(rv).strip()
        inv = ""
        if inv_col is not None and inv_col < ncol:
            iv = raw.iat[r, inv_col]
            inv = "" if pd.isna(iv) else re.sub(r"\D", "", str(iv))
        orders.append({"수령인": recv, "송장": inv, "amount": amt})
    return {"orders": orders, "net": net, "total": total}


_MASTER_RECV_KW = ("수령자", "수령인명", "수령인", "받는분", "받는사람", "수취인")
_MASTER_DATE_KW = ("주문일자", "주문일시", "주문일", "주문날짜", "결제일")


def _master_recip_map(path: Path, sheet: str, year: int, month: int) -> dict[int, set[str]]:
    """정산엑셀관리(전체 정산용) 선택 시트에서 대상 월의 {일(day): 수령자명 집합} 을 만든다.

    첫/마지막 발주서 파일 보정에 쓴다. 정산엑셀은 헤더가 0행이 아닐 수 있어(0행에 합계 등)
    수령자/주문일자 키워드가 있는 행을 헤더로 자동 탐지한다. 수령자='수령자'(발주서 '수령인명'과 동일인).
    """
    raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
    nrow, ncol = raw.shape

    def cell(r, c):
        v = raw.iat[r, c]
        return "" if pd.isna(v) else str(v).strip()

    header_row = None
    header: list[str] = []
    for r in range(min(20, nrow)):
        vals = [cell(r, c) for c in range(ncol)]
        if any(any(kw in v for kw in _MASTER_RECV_KW) for v in vals):
            header_row, header = r, vals
            break
    if header_row is None:
        raise ValueError(f"시트 '{sheet}'에서 수령자 헤더 행을 못 찾음")

    def col_of(kws):
        for i, h in enumerate(header):
            for kw in kws:
                if kw in h:
                    return i
        return None

    recv_col = col_of(_MASTER_RECV_KW)
    date_col = col_of(_MASTER_DATE_KW)
    if recv_col is None or date_col is None:
        raise ValueError(f"시트 '{sheet}'에서 수령자/주문일자 열을 못 찾음 (헤더: {header[:15]})")

    prefix = f"{year}{month:02d}"
    out: dict[int, set[str]] = {}
    for r in range(header_row + 1, nrow):
        d = "".join(ch for ch in cell(r, date_col) if ch.isdigit())[:8]
        if len(d) == 8 and d.startswith(prefix):
            name = cell(r, recv_col)
            if name:
                out.setdefault(int(d[6:8]), set()).add(name)
    return out


def _recips_in_days(recip_map: dict[int, set[str]], days) -> set[str]:
    """일(day) 목록에 해당하는 수령자명들의 합집합."""
    out: set[str] = set()
    for d in days:
        out |= recip_map.get(d, set())
    return out


@app.get("/api/aejulnun/master-sheets")
async def aejulnun_master_sheets():
    """정산엑셀관리 '전체 정산용 엑셀'(total)의 시트명 목록을 반환한다."""
    p = settlement_file_path("total")
    if not p:
        return {"sheets": [], "error": "정산엑셀관리에서 '전체 정산용 엑셀'을 먼저 업로드하세요."}
    try:
        wb = openpyxl.load_workbook(p, read_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        return {"sheets": sheets}
    except Exception as e:
        return {"sheets": [], "error": f"시트 목록 읽기 실패: {e}"}


@app.post("/api/aejulnun/calc")
async def calc_aejulnun(request: Request):
    """작업폴더 발주서들을 파일별로 계산해 [{파일명, 매입금, 에러}] + 합계를 반환한다.

    body: {year, month, sheet}. 중간 날짜 파일은 요약행 합계(부가세 포함)를 그대로,
    첫/마지막 날짜 파일은 정산엑셀(선택 시트) 수령인 매칭으로 보정해 순액→합계 재계산.
    """
    import calendar
    body = await request.json()
    try:
        year = int(body.get("year"))
        month = int(body.get("month"))
    except (TypeError, ValueError):
        return {"error": "계산할 년/월을 선택하세요.", "results": [], "total": None}
    sheet = (body.get("sheet") or "").strip()

    files = sorted(aejulnun_files(), key=lambda f: _leading_mmdd(f.name))
    if not files:
        return {"error": "폴더에 발주서 엑셀이 없습니다.", "results": [], "total": None}

    first, last = files[0], files[-1]
    last_month_day = calendar.monthrange(year, month)[1]

    def _file_day(f):
        """파일명 MMDD 에서 '일'(day)만. 그 파일의 월이 대상월이 아니면 None."""
        mmdd = _leading_mmdd(f.name)
        fm, fd = mmdd // 100, mmdd % 100
        return fd if fm == month else None

    # 첫/마지막 파일 보정용 수령자 집합 (실제 파일 날짜 기준 윈도우).
    #  - 첫 파일: 대상월 1일 ~ '첫 파일 날짜' 주문자 (앞의 전달분 제거)
    #  - 마지막 파일: '마지막 직전 파일 날짜' ~ 대상월 말일 주문자 (뒤의 다음달분 제거)
    total_path = settlement_file_path("total")
    first_set = last_set = None
    master_err = ""
    if not total_path:
        master_err = "정산엑셀관리에 '전체 정산용 엑셀' 미업로드 → 첫/마지막 보정 불가"
    elif not sheet:
        master_err = "정산엑셀 시트 미선택 → 첫/마지막 보정 불가"
    else:
        try:
            rmap = _master_recip_map(total_path, sheet, year, month)
            first_hi = _file_day(first) or 1
            first_set = _recips_in_days(rmap, range(1, first_hi + 1))
            last_lo = (_file_day(files[-2]) if len(files) >= 2 else None) or 1
            last_set = _recips_in_days(rmap, range(last_lo, last_month_day + 1))
        except Exception as e:
            master_err = f"정산엑셀 읽기 실패: {e}"

    # 상세보기용 윈도우 문자열 (첫/마지막)
    first_window = f"{month}/1 ~ {month}/{first_hi}" if first_set is not None else ""
    last_window = f"{month}/{last_lo} ~ {month}/{last_month_day}" if last_set is not None else ""

    def _row_name(o):
        return o["수령인"] if o["수령인"] else "반품"

    results: list[dict[str, Any]] = []
    total = 0
    for f in files:
        err = ""
        detail = None
        try:
            parsed = _parse_order_form(f)
        except Exception as e:
            results.append({"name": f.name, "amount": None, "error": f"파싱 실패: {e}", "detail": None})
            continue

        is_edge = (f == first) or (f == last)
        if is_edge:
            calc_net = sum(o["amount"] for o in parsed["orders"])
            if parsed["net"] is not None and calc_net != parsed["net"]:
                err = f"요약행순액 {parsed['net']:,} ≠ 계산순액 {calc_net:,}"
            target_set = first_set if f == first else last_set
            if target_set is None:
                # 보정 불가 → 요약행 합계로 대체하고 사유를 에러열에 남김
                amount = parsed["total"] if parsed["total"] is not None else (calc_net + round(calc_net * 0.1))
                err = (err + " / " if err else "") + master_err
            else:
                # 이번 달 주문(수령자)만 남김. 송장 없는 '반품' 행은 유지.
                kept = [o for o in parsed["orders"] if (not o["수령인"]) or (o["수령인"] in target_set)]
                removed = [o for o in parsed["orders"] if o["수령인"] and o["수령인"] not in target_set]
                net = sum(o["amount"] for o in kept)
                vat = round(net * 0.1)
                amount = net + vat
                detail = {
                    "edge": "first" if f == first else "last",
                    "window": first_window if f == first else last_window,
                    "full_net": calc_net,
                    "kept": [{"name": _row_name(o), "amount": o["amount"]} for o in kept],
                    "removed": [{"name": _row_name(o), "amount": o["amount"]} for o in removed],
                    "net": net,
                    "vat": vat,
                }
        else:
            amount = parsed["total"]
            if amount is None:
                err = "요약행(합계) 못 찾음"
                amount = 0

        results.append({"name": f.name, "amount": amount, "error": err, "detail": detail})
        total += amount or 0

    print(f"[에이준줄눈] 계산: {year}년 {month}월, {len(results)}개 파일, 합계 {total:,}원", flush=True)
    send_log(f"[에이준줄눈] 계산: {year}년 {month}월 → {total:,}원 ({len(results)}개)")
    return {
        "results": results,
        "total": total,
        "month": f"{year}-{month:02d}",
        "sheet": sheet,
        "master_note": master_err,
        "pending": False,
    }


@app.delete("/api/aejulnun")
async def delete_aejulnun():
    """에이준줄눈 작업폴더를 비우고 메타를 삭제한다."""
    _wipe_aejulnun_dir()
    if AEJULNUN_META.exists():
        try:
            AEJULNUN_META.unlink()
        except Exception:
            pass
    print("[에이준줄눈] 작업폴더 비움", flush=True)
    send_log("[에이준줄눈] 작업폴더 비움")
    return {"ok": True, **_aejulnun_status()}


# ── 수집 API ──────────────────────────────────────────────
@app.post("/api/collect")
async def start_collect(request: Request):
    """매입금 수집을 시작한다."""
    global _collect_running
    if _collect_running:
        return {"error": "이미 수집이 진행 중입니다"}

    body = await request.json()
    sites: list[dict[str, str]] = body.get("sites", [])
    start_date: str = body.get("start_date", "")
    end_date: str = body.get("end_date", "")

    if not sites or not start_date or not end_date:
        return {"error": "사이트 목록과 날짜를 입력해주세요"}

    # start_date에서 year, month 추출
    parts = start_date.split("-")
    year, month = int(parts[0]), int(parts[1])

    _collect_running = True
    _collect_status.clear()
    _logs.clear()  # 새 수집 시작 시 로그 버퍼 초기화

    thread = threading.Thread(
        target=_run_all_sites, args=(sites, year, month), daemon=True
    )
    thread.start()
    return {"ok": True, "message": "수집을 시작했습니다"}


@app.get("/api/collect/status")
async def collect_status(since: int = 0):
    """현재 수집 진행상황 + since 인덱스 이후의 새 로그를 반환한다."""
    new_logs = _logs[since:] if since < len(_logs) else []
    return {
        "running": _collect_running,
        "sites": _collect_status,
        "logs": new_logs,
        "log_count": len(_logs),
    }


# ── WebSocket ─────────────────────────────────────────────
@app.websocket("/ws/logs")
async def ws_logs(websocket: WebSocket):
    """실시간 로그 WebSocket."""
    await websocket.accept()
    _ws_clients.add(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        _ws_clients.discard(websocket)


# ── 실행 ──────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.server:app", host="0.0.0.0", port=8000, reload=True)
