"""'그때그때' 탭 - 페브스토어 판매가/매입가 비교용 엑셀 파싱.

스마트스토어 "상품관리" 다운로드 포맷(도매처별 시트)에서 판매자상품코드/판매상태/
판매가/기본배송비를 읽는다. 시트마다 컬럼 개수·순서가 달라(예: 오너클랜 시트는
'판매자상품코드2' 컬럼이 끼어 있어 뒤 컬럼이 하나씩 밀림) 컬럼명 키워드로 찾는다.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


def _s(value) -> str:
    """셀 값을 표시용 문자열로. NaN/None 은 빈 문자열."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _to_int(value) -> int | None:
    """'35,800' / '35,800원' / 35800.0 등을 정수로. 빈 값은 None."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    s = re.sub(r"[^0-9.\-]", "", str(value))
    if s in ("", "-", ".", "-."):
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def _find_col(df: pd.DataFrame, keywords: list[str]) -> str | None:
    """컬럼명에 keywords 중 하나라도 포함된 첫 컬럼명을 반환. 없으면 None."""
    for col in df.columns:
        name = str(col)
        for kw in keywords:
            if kw in name:
                return col
    return None


def _match_sheet(site_name: str, sheet_names: list[str]) -> str | None:
    """사이트 한글명과 시트명을 양방향 부분일치로 매칭한다.

    실제 업로드 파일의 시트명은 사이트 전체명과 다를 수 있다
    (예: 시트 'JTC' ↔ 사이트명 'JTC코리아', 시트 '히트' ↔ '히트가구').
    """
    for sheet in sheet_names:
        if sheet == site_name or sheet in site_name or site_name in sheet:
            return sheet
    return None


def parse_febstore(path: Path, site_list: list[dict[str, str]]) -> list[dict[str, Any]]:
    """업로드된 엑셀에서 도매처별 시트를 찾아 상품 목록을 읽는다.

    반환: [{"site_name","slug","code","status","price","ship",
            "cost","cost_ship","ship_note","note"}]
    cost/cost_ship 은 아직 매입가 조회 전이라 None (오너클랜은 조회 버튼으로 채움),
    ownerclan 이 아닌 slug 는 note="미구현" 으로 고정한다.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()

    rows: list[dict[str, Any]] = []
    for site in site_list:
        name, slug = site["name"], site["slug"]
        sheet = _match_sheet(name, sheet_names)
        if sheet is None:
            continue

        df = pd.read_excel(path, sheet_name=sheet, dtype=object)
        code_col = _find_col(df, ["판매자상품코드"])
        status_col = _find_col(df, ["판매상태"])
        price_col = _find_col(df, ["판매가"])
        ship_col = _find_col(df, ["기본배송비"])
        if code_col is None:
            continue

        for _, r in df.iterrows():
            code = _s(r.get(code_col))
            if not code:
                continue
            rows.append({
                "site_name": name,
                "slug": slug,
                "code": code,
                "status": _s(r.get(status_col)) if status_col else "",
                "price": _to_int(r.get(price_col)) if price_col else None,
                "ship": _to_int(r.get(ship_col)) if ship_col else None,
                "cost": None,
                "cost_ship": None,
                "ship_note": "",
                "note": "" if slug == "ownerclan" else "미구현",
            })
    return rows


def has_any_site_sheet(path: Path, site_list: list[dict[str, str]]) -> bool:
    """업로드 파일에 도매처 시트가 하나라도 있는지 확인한다(업로드 검증용)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
    except Exception:
        return False
    return any(_match_sheet(site["name"], sheet_names) for site in site_list)


def ownerclan_codes(rows: list[dict[str, Any]]) -> list[str]:
    """rows 에서 오너클랜 상품코드만 뽑는다(OWNER_ 접두어는 떼고 사이트 자체 코드로).

    페브스토어 엑셀의 판매자상품코드는 'OWNER_W6D9C27' 형태이고, 오너클랜
    상품상세 URL 은 접두어 없는 자체 코드('W6D9C27')를 쓴다.
    """
    codes = []
    for r in rows:
        if r["slug"] != "ownerclan":
            continue
        code = r["code"]
        if code.upper().startswith("OWNER_"):
            code = code[6:]
        elif code.upper().startswith("OWNER"):
            code = code[5:]
        codes.append(code)
    return codes


def build_febstore_excel(rows: list[dict[str, Any]]) -> bytes:
    """페브스토어 비교표를 엑셀 바이트로 만든다."""
    import io

    def ship_display(r: dict[str, Any]) -> Any:
        if r["slug"] != "ownerclan":
            return "미구현"
        if r.get("cost_ship") is not None:
            return "무료" if r["cost_ship"] == 0 else r["cost_ship"]
        return r.get("ship_note") or ""

    detail = pd.DataFrame([
        {
            "판매자관리코드": r["code"],
            "도매처": r["site_name"],
            "판매상태": r["status"],
            "판매가": r["price"],
            "배송비": "무료" if r.get("ship") == 0 else r.get("ship"),
            "매입가": r["cost"] if r["slug"] == "ownerclan" else "미구현",
            "매입배송비": ship_display(r),
            "비고": r.get("note", ""),
        }
        for r in rows
    ])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="판매가매입가비교", index=False)
    return buf.getvalue()
